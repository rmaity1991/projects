"""
Created on Friday June 4 2021

@author: Emeka Ene
"""

import os

pre = 'start /wait cmd /k'

'''
class CIN:
def __init__(self):
def generator(self, AS400_info):
pass
def manual_gen(self)
def
'''

class excel:
    
    def __init__(self, Comm_Type, COM_Port, Baud_Rate, Bits, Parity, StopBits, Address):
        self.Comm_Type = Comm_Type
        self.COM_Port = COM_Port
        self.Baud_Rate = Baud_Rate
        self.Bits = Bits
        self.Parity = Parity
        self.StopBits = StopBits
        self.Address = Address
     
    def sparkly_connection_info(self):
        command = "--connection " + self.Comm_Type + ',' + self.COM_Port + ',' + self.Baud_Rate + self.Bits + self.Parity + self.StopBits + ',' + self.Address
        return command

    # sparklyread('Serial','COM3','57600','8','N','2','197')
    def sparkly_read_DeviceInfo(self):
        command = "sparkly device read-data " + self.sparkly_connection_info()
        return os.system(command)


    def excel_import_write(self):
        from openpyxl import load_workbook

        # Location of the excel File
        # Open Workbook
        wb = load_workbook(filename = 'Z:\\@@@Controls\\1_SW&DiagDevelopment\\3-Step-Process\\Parameter_Ownership_Overall_Rev1.xlsx', read_only=True)
        ws = wb['Main']

        #print(ws.title)

        # Get Column number for excel sheet customer (Dollar Tree, Publix or Target)
        col_customer = ''
        col_acronym = ''
        #customer = input('What customer parameters do you need (Dollar Tree, Publix or Target):\n')
        customer = 'Publix'
        for cell in ws[1]:
            #print(cell.value)
            if cell.value == customer:
                col_customer = cell.column - 1
                break
            if cell.value == 'Acronym':
                col_acronym = cell.column - 1

        #print(col_acronym)

        # Read and write data using Sparkly into controller
        final_data =''
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            acronym_cell = row[col_acronym]
            customer_cell = row[col_customer]

            if acronym_cell is not None and customer_cell is not None:
                acronym_value = acronym_cell.value
                customer_value = customer_cell.value
                if acronym_value is not None and customer_value is not None:
                    if final_data == '':
                        final_data = acronym_value +"=" + customer_value
                    else:
                        final_data = final_data + ' ' + acronym_value + "=" + customer_value

        print(final_data)
        output = pre + 'sparkly parameters write --parameter-list ' +'"' + final_data +'" ' + self.sparkly_connection_info()
        #print(output.split(" "))
        return os.system(output)
    def sparklyread_read_data(self,data):
        final_data =""
        for vals in data:
            if final_data =="":
                final_data= final_data + vals
            else:
                final_data = final_data +" " + vals
        output = pre + 'sparkly parameters read --parameter-list ' +'"' + final_data +'" ' + self.sparkly_connection_info()
        #print(output.split(" "))
        return os.system(output)

    def excel_import_write_2(self):
        from openpyxl import load_workbook

        # Location of the excel File
        # Open Workbook
        wb = load_workbook(filename = 'Parameter_Ownership_Overall_Rev1.xlsx', read_only=True)
        ws = wb['Main']

        #print(ws.title)

        # Get Column number for excel sheet customer (Dollar Tree, Publix or Target)
        col_customer = ''
        col_acronym = ''
        #customer = input('What customer parameters do you need (Dollar Tree, Publix or Target):\n')
        customer = 'Aldi'
        for cell in ws[1]:
            #print(cell.value)
            if cell.value == customer:
                col_customer = cell.column - 1
                break
            if cell.value == 'Acronym':
                col_acronym = cell.column - 1

        #print(col_acronym)
        list_short_name = []
        list_cust_val = []
        # Read and write data using Sparkly into controller
        final_data =''
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            acronym_cell = row[col_acronym]
            customer_cell = row[col_customer]

            if acronym_cell is not None and customer_cell is not None:
                acronym_value = acronym_cell.value
                customer_value = customer_cell.value
                
                list_short_name.append(acronym_value)
                list_cust_val.append(customer_value)
                if acronym_value is not None and customer_value is not None:
                    if final_data == '':
                        final_data = acronym_value +"=" + customer_value
                    else:
                        final_data = final_data + ' ' + acronym_value + "=" + customer_value
                        
        #list_short_name.remove(" None")
        list_short_name = list(filter((None).__ne__, list_short_name))


        #print(final_data)
        #output = pre + 'sparkly parameters write --parameter-list ' +'"' + final_data +'" ' + self.sparkly_connection_info()
        #print(output.split(" "))
        return list_short_name

testing = excel('Serial','COM9','19200','8','N','2','24')


#print(testing.sparkly_connection_info())
#print(testing.sparkly_read_DeviceInfo())
#print(testing.excel_import_write())


out_1=testing.excel_import_write_2()# -*- coding: utf-8 -*-

print(out_1)
"""
print(testing.sparkly_connection_info())
print(testing.sparkly_read_DeviceInfo())

print(testing.sparklyread_read_data(out_1))
#print(testing.excel_import_write())

"""

"""
Spyder Editor

This is a temporary script file.
"""

