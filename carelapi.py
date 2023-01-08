"""
    Initial Release : Rohit Maity
    Date : 21/10/2022
"""


from pymodbus.client.sync import ModbusSerialClient,ModbusTcpClient
import os
import json
import datetime,time
import openpyxl


class Utility:

    def __init__(self) -> None:
        pass
    @classmethod
    def ton(cls,enable,seconds)->bool:
        if enable:
            if(seconds<0):
                return False
            else:
                time.sleep(seconds)
                enable=False
                return True
        else:
            return False
    @classmethod
    def toff(cls,enable,seconds)->bool:
        if enable:
            if(seconds<0):
                return True
            else:
                time.sleep(seconds)
                enable=False
                return False
        else:
            return True
 
class Carel:

    controllers={
        1:{
            'name':'mpxone',
            'comm':'Serial'
        },

        2:{
            'name':'mpxone',
            'comm':'Modbus'
        },

        3:{
            'name':'cpco',
            'comm':'Serial'
        },

        4:{
            'name':'cpco',
            'comm':'modbus'
        },
        5:{
            'name':'pr300',
            'comm':'modbus'
        },
        6:{
            'name':'pr300',
            'comm':'serial'
        }
    }

    def __init__(self,controller_type:int):

        self.connection=None
        self.result=""
        self.wb=openpyxl.Workbook()
        self.sheet=self.wb['Sheet']
        self.wb.remove(self.sheet)
        self.wb.create_sheet("Results")
        self.sheet=self.wb['Results']
        self.sheet.cell(row=1,column=1).value="Date"
        self.sheet.cell(row=1,column=2).value="Controller"
        self.sheet.cell(row=1,column=3).value="Test Name"
        self.sheet.cell(row=1,column=4).value="Status"

        if controller_type in Carel.controllers:
            self.controller=controller_type

    def sparklyConnection(self, COM_Port='COM11', Baud_Rate="19200", Bits='8', Parity='N',StopBits='2', Address='199'):

        """
        Define the following paramters for sparkly connection (Values are Examples):

        1) COM_Port : COM11 
        2) Baud_Rate : 19200
        3) Bits : 8
        4) Parity : N
        5) StopBits :2
        6) Address : 199
        """
        self.COM_Port = COM_Port
        self.Baud_Rate = Baud_Rate
        self.Bits =Bits
        self.Parity = Parity
        self.StopBits = StopBits
        self.Address =Address

        self.connection ="--connection " + Carel.controllers[1]['comm'] +','+ self.COM_Port+','+self.Baud_Rate + self.Bits + self.Parity +self.StopBits +','+self.Address


    def modbusConnection(self):
        pass
    
    def serialConnection(self):
        pass

    def _connect(self):
        print("\t \t \t \t \t Connection Succesfull \t \t \t \t \t")
        if(self.controller==1):
            print("\t \t \t \t \t Enter the COM Port to be used \t \t \t \t \t")
            comm=str(input("\t \t \t \t \t \t"))
            print("\t \t \t \t \t Enter the Device Address to be used \t \t \t \t \t")
            device=str(input("\t \t \t \t \t \t"))
            self.sparklyConnection(COM_Port=comm,Address=device)
            command = "sparkly device read-data " + self.connection
            os.system(command)
        
        if(self.controller==2):
            pass

        if(self.controller==3):
            pass

        if(self.controller==4):
            pass



    def _lightTest(self):
        
        if(self.connection==None):
            print("\t \t \t \t \t First test the connection using option 1 to create client to the controller \t \t \t \t \t")
        else:
            if(self.controller==1):
               print("\t \t \t \t \t Light Test Being Performed \t \t \t \t \t")
               message="{0}={1}".format('Lht',1)
               command="sparkly parameters write --parameter-list "+ " " + message + " " +self.connection
               os.system(command)


            if(self.controller==2):
                pass

            if(self.controller==3):
                pass

            if(self.controller==4):
                pass

            result=[]
            result.append(str(datetime.datetime.now()))
            result.append(Carel.controllers[self.controller]['name'])
            result.append('Light Test')
            result.append('Passed')

            self.sheet.append(result)

            self.wb.save("Test Results.xlsx")

               



    def _eevTest(self):
        print("\t \t \t \t \t EEV Test Being Performed \t \t \t \t \t")

        if(self.controller==2):
            pass

        if(self.controller==3):
            pass

        if(self.controller==4):
            pass


    def __str__(self) -> str:
        return "\t \t \t \t \t The object has been created, start performing tests in the option 2 of main menu \t \t \t \t \t"

    def __del__(self):
        print("\t \t \t \t \t The object has been deleted, redefine controller in main menu \t \t \t \t \t")





